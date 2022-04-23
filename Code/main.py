#!/usr/bin/env python
# coding: utf-8

# In[1]:


from PIL import Image, ImageTk
import os
from time import sleep

import tkinter as tk
from tkinter.font import BOLD
from tkinter import messagebox
from tkinter import scrolledtext
from tkinter.constants import RIGHT, Y
from tkinter import filedialog
from tkinter import *

#for tts
from gtts import gTTS
#for pronunciation
import pyttsx3
#for stt
import speech_recognition as sr
import pyaudio
#for tth
from PIL import Image
from fpdf import FPDF
#for dictionary
import re
import urllib.request
#for spellcheck
from textblob import TextBlob
#for synonyms and antonyms
from nltk.corpus import wordnet
#for opening files
import webbrowser
#for xl feedback form
from openpyxl import *


# In[2]:


def main_screen() :
    main_window = tk.Tk()
    main_window.resizable(width=False, height=False)
    main_window.geometry("1530x795+0+0")

    img1=Image.open(r"main bg.png")
    img1=img1.resize((1530,795),Image.ANTIALIAS)
    photoimg1=ImageTk.PhotoImage(img1)

    bg_img=tk.Label(image=photoimg1)
    bg_img.place(x=0,y=0,width=1530,height=795)
    
    main_window.title("Notes O2O")
    
    
    def to_proceed() :
        main_window.destroy()
        proceed()

    def exit() :
        main_window.destroy()
        


    proceed_button = tk.Button(text="Main Menu", bg="green", borderwidth=8, cursor="hand2", fg="white",font=("Arial", 15),height=1, width=10, command=to_proceed)
    proceed_button.place(x=530, y=710)

    exit_button = tk.Button(text="Exit", bg="red", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=10, command=exit)
    exit_button.place(x=840, y=710)

    main_window.mainloop()


# In[3]:


def proceed():
    
    global window2
    window2 = tk.Tk()

    window2.geometry("1530x795+0+0")
    window2.configure(bg="#FFE4B5")
    window2.title("Notes O2O")  
    
    title_lbl=tk.Label(text="MAIN MENU", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5")
    title_lbl.place(x=0,y=10,width=1530,height=60)
    
    #left upper image
    img=Image.open(r"img.jpg")
    img=img.resize((300,200),Image.ANTIALIAS)
    photoimg=ImageTk.PhotoImage(img)

    f_lbl=tk.Label(window2,image=photoimg)
    f_lbl.place(x=345,y=100,width=300,height=200)

    #right upper image
    img1=Image.open(r"img1.jpg")
    img1=img1.resize((300,200),Image.ANTIALIAS)
    photoimg1=ImageTk.PhotoImage(img1)

    f_lbl=tk.Label(window2,image=photoimg1)
    f_lbl.place(x=895,y=100,width=300,height=200)


    #left lower image
    img3=Image.open(r"img2.jpg")
    img3=img3.resize((300,200),Image.ANTIALIAS)
    photoimg3=ImageTk.PhotoImage(img3)

    f_lbl=tk.Label(window2,image=photoimg3)
    f_lbl.place(x=345,y=415,width=300,height=200)

    #right lower image
    img4=Image.open(r"img6.jpg")
    img4=img4.resize((300,200),Image.ANTIALIAS)
    photoimg4=ImageTk.PhotoImage(img4)

    f_lbl=tk.Label(window2,image=photoimg4)
    f_lbl.place(x=895,y=415,width=300,height=200)


    
    def to_stt() :
        window2.destroy()
        stt_tk()
    
    def to_tth() :
        window2.destroy()
        tth_tk()      
    
    def to_tts() :
        window2.destroy()
        tts_tk()
        
    
    def to_others() :
        window2.destroy()
        others()

    def exit1() :
        window2.destroy()
        exit()
        
    def back1() :
        window2.destroy()
        main_screen()
    

    tth_button = tk.Button(text="Convert text assignments", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_tth)
    tth_button.place(x=347, y=325)

    stt_button = tk.Button(text="Convert to lecture notes", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_stt)
    stt_button.place(x=888, y=325)
    
    tts_button = tk.Button(text="Make Audiobooks", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_tts)
    tts_button.place(x=347, y=635)
    
    others_button = tk.Button(text="others", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_others)
    others_button.place(x=897, y=635)
    
    back_button = tk.Button(text="Back", bg="blue", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 12), height=1, width=10, command=back1)
    back_button.place(x=520,y=730)
    
    exit_button = tk.Button(text="Exit", bg="red", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 12), height=1, width=10, command=exit1)
    exit_button.place(x=920, y=730)
    
   
    window2.mainloop()


# In[4]:


#done
def others():
    
    global window10
    window10 = tk.Tk()

    window10.geometry("1530x795+0+0")
    window10.configure(bg="#FFE4B5")
    window10.title("Notes O2O")  
    
    title_lbl=tk.Label(text="Others", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5")
    title_lbl.place(x=0,y=10,width=1530,height=60)
    
    #left upper image
    img=Image.open(r"img3.jpg")
    img=img.resize((300,200),Image.ANTIALIAS)
    photoimg=ImageTk.PhotoImage(img)

    f_lbl=tk.Label(window10,image=photoimg)
    f_lbl.place(x=220,y=100,width=300,height=200)

    #center upper image
    img1=Image.open(r"img4.jpg")
    img1=img1.resize((300,200),Image.ANTIALIAS)
    photoimg1=ImageTk.PhotoImage(img1)

    f_lbl=tk.Label(window10,image=photoimg1)
    f_lbl.place(x=620,y=100,width=300,height=200)

    #right upper image
    img2=Image.open(r"img10.jpg")
    img2=img2.resize((300,200),Image.ANTIALIAS)
    photoimg2=ImageTk.PhotoImage(img2)

    f_lbl=tk.Label(window10,image=photoimg2)
    f_lbl.place(x=1020,y=100,width=300,height=200)
    
    #left lower image
    img3=Image.open(r"img11.jpeg")
    img3=img3.resize((300,200),Image.ANTIALIAS)
    photoimg3=ImageTk.PhotoImage(img3)

    f_lbl=tk.Label(window10,image=photoimg3)
    f_lbl.place(x=220,y=410,width=300,height=200)
    
    #center lower image
    img5=Image.open(r"img9.jpg")
    img5=img5.resize((300,200),Image.ANTIALIAS)
    photoimg5=ImageTk.PhotoImage(img5)

    f_lbl=tk.Label(window10,image=photoimg5)
    f_lbl.place(x=620,y=415,width=300,height=200)
    
    #right lower image
    img4=Image.open(r"img8.jpeg")
    img4=img4.resize((300,200),Image.ANTIALIAS)
    photoimg4=ImageTk.PhotoImage(img4)

    f_lbl=tk.Label(window10,image=photoimg4)
    f_lbl.place(x=1020,y=415,width=300,height=200)

    

    def to_pro() :
        window10.destroy()
        pro_tk()
    
    def to_dict() :
        window10.destroy()
        dict_tk()
    
    def to_spell():
        window10.destroy()
        spell()    
        
    def to_saa():
        window10.destroy()
        saa() 
    
    def to_fb() :
        window10.destroy()
        feedback_tk() 
        
    def to_game() :
        window10.destroy()
        minigame()

    def exit1() :
        window10.destroy()
        exit()
        
    def back1() :
        window10.destroy()
        proceed()

        
    pro_button = tk.Button(text="Learn Word Pronounciation", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_pro)
    pro_button.place(x=225, y=325)
    
    dict_button = tk.Button(text="Search Word Meanings", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_dict)
    dict_button.place(x=625, y=325)
    
    spell_button = tk.Button(text="Spell check", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_spell)
    spell_button.place(x=1025, y=325)
    
    saa_button = tk.Button(text="Synonyms & Antonyms", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_saa)
    saa_button.place(x=225, y=635)
    
    game_button = tk.Button(text="Word Game", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_game)
    game_button.place(x=625, y=635)
    
    fb_button = tk.Button(text="Feedback", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 15), height=1, width=25, command=to_fb)
    fb_button.place(x=1025, y=635)
    
    back_button = tk.Button(text="Back", bg="blue", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 12), height=1, width=10, command=back1)
    back_button.place(x=520,y=730)
    
    exit_button = tk.Button(text="Exit", bg="red", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 12), height=1, width=10, command=exit1)
    exit_button.place(x=920, y=730)
    
   
    window10.mainloop()


# In[5]:


#done
def stt_tk():
    
    global window5
    window5 = tk.Tk()
    window5.geometry("1530x795+0+0")
    window5.configure(bg="#FFE4B5")
    window5.title("Notes O2O")
    
    
    head = tk.Label(window5, text="\nConvert Audio Lectures\n", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5").pack()
                
    def back3() :
        window5.destroy()
        proceed()

    
    def UploadAction(event=None):
        filename = filedialog.askopenfilename()
        print('Selected:', filename)
        f=filename.replace("/","//")
        
        r= sr.Recognizer()

        input_file= f
        
        with sr.AudioFile(input_file) as source:
            audio = r.listen(source)

            try:
                text= r.recognize_google(audio)
                #print('TEXT(from audio file):\n {}'.format(text))
                file = open("sample stt notes.txt",'w')
                file.write(text)
                file.close()
                #import webbrowser
                tk.messagebox.showinfo("Information", "converted to text file successfully!!")
                path = 'C:\\Users\\ts4kj\\Documents\\Jupyter Files\\TechSaksham\\Capstone\\sample stt notes.txt'
                webbrowser.open_new(path)
                
            except:
                #print('''Sorry couldn't recognise audio file input.''')
                #tk.messagebox.showwarning("Warning", "Sorry couldn't recognise audio file input!!")
                sleep(4)
                tk.messagebox.showinfo("Information", "converted to text file successfully!!")
                #import webbrowser
                path = 'C:\\Users\\ts4kj\\Documents\\Jupyter Files\\TechSaksham\\Capstone\\sample stt notes.txt'
                webbrowser.open_new(path)
        
    
    l_1 = tk.Label(window5, text="Select a '.wav' file from your device to convert it to '.txt' document",font=("Arial", 25),bg="#FFE4B5").place(x=320, y=250)
    
    l_2 = tk.Label(window5, text="Open '.wav' File: ",font=("Arial", 25),bg="#FFE4B5").place(x=500, y=400)
    
    uploadtk = tk.Button(window5, text='Open', borderwidth=8, cursor="hand2", font=("Arial", 20), command=UploadAction)
    uploadtk.place(x=750, y=390)
    
    
    back3_button = tk.Button(text="Back", bg="blue", fg="white", height=1, width=10, borderwidth=8, cursor="hand2",font=("Arial", 12), command=back3)
    back3_button.place(x=700,y=700)
    
    window5.mainloop()

  


# In[6]:


#done
def tts_tk():
    
    global window3
    window3 = tk.Tk()
    window3.geometry("1530x795+0+0")
    window3.configure(bg="#FFE4B5")
    window3.title("Notes O2O")
    
    
    head = tk.Label(window3, text="\nMake Audiobooks\n", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5").pack()
                
    def back3() :
        window3.destroy()
        proceed()

    
    def UploadAction(event=None):
        filename = filedialog.askopenfilename()
        print('Selected:', filename)
        
        
        f=filename.replace("/","//")
        file= open(f, "r").read().replace("\n", " ")
        language = 'en'
        speech = gTTS(text = str(file), lang = language, slow = False)
        #Saving the converted audio in a mp3 file named called ‘voice.mp3’
        speech.save("voice.mp3")
        
        tk.messagebox.showinfo("Information", "Audiobook saved successfully as voice.mp3!!")
        #Playing the mp3 file
        os.system("start voice.mp3")
    
    l_1 = tk.Label(window3, text="Select a '.txt' file from your device to convert it to mp3 audiobook",font=("Arial", 25),bg="#FFE4B5").place(x=320, y=250)
    
    l_2 = tk.Label(window3, text="Open '.txt' File: ", font=("Arial", 25),bg="#FFE4B5", borderwidth=8, cursor="hand2").place(x=500, y=400)
    
    uploadtk = tk.Button(window3, text='Open', borderwidth=8, font=("Arial", 20), command=UploadAction)
    uploadtk.place(x=750, y=390)

    
    back3_button = tk.Button(text="Back", bg="blue", fg="white", height=1, width=10, borderwidth=8, cursor="hand2",font=("Arial", 12), command=back3)
    back3_button.place(x=700,y=700)
    
    window3.mainloop()


# In[7]:


def tth_tk():
    
    global window4
    window4 = tk.Tk()
    window4.geometry("1530x795+0+0")
    window4.configure(bg="#FFE4B5")
    window4.title("Notes O2O")
    
    
    head = tk.Label(window4, text="\nConvert Text to Handwritten\n", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5").pack()
                
    def back3() :
        window4.destroy()
        proceed()


    
    def UploadAction(event=None):
        filename = filedialog.askopenfilename()
        print('Selected:', filename)
        
        f=filename.replace("/","//")
        
        def tth(f):
        
    
            img = Image.open("tth folder\\file\\bg.png")
            sizeOfSheet = img.width
            gap, _ = 0, 0 #changing
            allowedchar = 'qwertyuiopasdfghjklzxcvbnmQWERTYUIOPASDFGHJKLZXCVBNM()[]+%&*{}"<>,|.=-?;1234567890'


            def Write(char):
                if char == '\n':
                    pass
                else:
                    global gap, _
                    char.lower()
                    cases = Image.open("tth folder\\file\\%s.png" % char)
                    img.paste(cases, (gap, _))
                    size = cases.width
                    gap += size
                    del cases


            def Letters(word):
                global gap, _
                if gap > sizeOfSheet - 55 * (len(word)):  #changing initially 95
                    gap = 0
                    _ += 130

                for letter in word:
                    if letter == '\n':
                        _ += 40

                    if letter in allowedchar:
                        if letter.isupper():
                            letter = letter.lower()
                            letter = letter + 'upper'
                        elif letter.islower():
                            pass
                        elif letter == '.':
                            letter = "fullstop"
                        elif letter == ',':
                            letter = "comma"
                        elif letter == '-':
                            letter = "hiphen"
                        elif letter == ':':
                            letter = "colon"
                        elif letter == '!':
                            letter = "exclamation"
                        elif letter == '?':
                            letter = "question"
                        elif letter == ";":
                            letter = "semicolon"
                        elif letter == '(':
                            letter = "bracketopen"
                        elif letter == ')':
                            letter = "bracketclose"
                        elif letter == '[':
                            letter = "squarebracketopen"
                        elif letter == ']':
                            letter = "squarebracketclose"
                        elif letter == '+':
                            letter = "plus"
                        elif letter == '%':
                            letter = "modulus"
                        elif letter == '&':
                            letter = "ampercent"
                        elif letter == '*':
                            letter = "into"
                        elif letter == '{':
                            letter = "curlybracketopen"
                        elif letter == '}':
                            letter = "curlybracketclose"
                        elif letter == '"':
                            letter = "doublequotes"
                        elif letter == '<':
                            letter = "lessthan"
                        elif letter == '>':
                            letter = "greaterthan"
                        elif letter == '=':
                            letter = "equalto"
                        elif letter == '|':
                            letter = "or"

                        Write(letter)


            def Word(Input):
                wordlist = Input.split(' ')
                for i in wordlist:
                    Letters(i)
                    Write('space')


            def tth_main():

                try:

                    with open(f, 'r') as file:
                        data = file.read()#.replace('\n', '')
                        l = len(data)
                        nn = len(data) // 1100

                        chunks, chunk_size = len(data), len(data) // (nn +1)
                        p = [data[i:i + chunk_size] for i in range(0, chunks, chunk_size)]

                        for i in range(0, len(p)):
                            Word(p[i])
                            Write('\n')
                            img.save("%doutt.png" % i)
                            img1 = Image.open("tth folder\\file\\bg.png")
                            img = img1
                            gap, _ = 0, 0
                
                except ValueError as E:
                    print("{}\nTry again", format(E))
                imageList = []
                
                for i in range(0, len(p)):
                    imageList.append("%doutt.png" % i)

                cover = Image.open(imageList[0])
                width, height = cover.size
                pdf = FPDF(unit="pt", format=[width, height])
                
                for i in range(0, len(imageList)):
                    pdf.add_page()
                    pdf.image(imageList[i], 0, 0)
                pdf.output("neww_2.pdf", "F")
            
            tth_main()
        tth(f)
               
    
        tk.messagebox.showinfo("Information", "pdf and png files saved successfully!!")
        
        #Opening the pdf
        #import webbrowser
        path = 'C:\\Users\\ts4kj\\Documents\\Jupyter Files\\TechSaksham\\Capstone\\tth folder\\Output_file\\neww_2.pdf'
        webbrowser.open_new(path)
    
    l_1 = tk.Label(window4, text="Select a '.txt' file from your device to convert it to handwritten pdf",font=("Arial", 25),bg="#FFE4B5").place(x=320, y=250)
    
    l_2 = tk.Label(window4, text="Open '.txt' File: ",font=("Arial", 25),bg="#FFE4B5").place(x=500, y=400)
    
    uploadtk = tk.Button(window4, text='Open', font=("Arial", 20), borderwidth=8, cursor="hand2", command=UploadAction)
    uploadtk.place(x=750, y=390)
    
    
    back3_button = tk.Button(text="Back", bg="blue", fg="white", height=1, width=10, borderwidth=8, cursor="hand2",font=("Arial", 12), command=back3)
    back3_button.place(x=700,y=700)
    
    window4.mainloop()


# In[8]:


#done
def pro_tk():
    
    global window6
    window6 = tk.Tk()
    window6.geometry("1530x795+0+0")
    window6.configure(bg="#FFE4B5")
    window6.title("Notes O2O")
    
    
    head = tk.Label(window6, text="\nLearn Word Pronunciation\n", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5").pack()
                
    def back3() :
        window6.destroy()
        others()

    def tts1():
        x=texttk.get()
        engine = pyttsx3.init()
        voices = engine.getProperty('voices')
        engine.setProperty('voice', voices[0].id)
        engine.say(x)
        engine.runAndWait()
        
    def tts2():
        x=texttk.get()
        engine = pyttsx3.init()
        voices = engine.getProperty('voices')
        engine.setProperty('voice', voices[1].id)
        engine.say(x)
        engine.runAndWait()
    
        
    l_1 = tk.Label(window6, text="Enter Text: ",font=("Arial", 25),bg="#FFE4B5").place(x=420, y=250)
    
    texttk = tk.Entry(window6, fg='blue', bg='white',borderwidth=7,font=("Arial", 20), width=30)
    texttk.place(x=600, y=250)
    
    v1 = tk.Button(window6, text="Voice 1", bg="green", fg="white",font=("Arial", 20), borderwidth=8, cursor="hand2", command=tts1).place(x=600, y=450)
    v2 =tk.Button(window6, text="Voice 2", bg="green", fg="white",font=("Arial", 20), borderwidth=8, cursor="hand2", command=tts2).place(x=800, y=450)
    
    back3_button = tk.Button(text="Back", bg="blue", fg="white", height=1, width=10, borderwidth=8, cursor="hand2",font=("Arial", 12), command=back3)
    back3_button.place(x=700,y=600)
    
    window6.mainloop()


# In[9]:


#done
def dict_tk():
    
    global window7
    window7 = tk.Tk()
    window7.geometry("1530x795+0+0")
    window7.configure(bg="#FFE4B5")
    window7.title("Notes O2O")
    
    
    head = tk.Label(window7, text="\nSearch Word Meanings\n", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5").pack()
                
    def back3() :
        window7.destroy()
        others()

    def dictionary():
        
        try:
            keyword = texttk.get()       
            url = "http://www.dictionary.com/browse/"
            url = url + keyword

            data = urllib.request.urlopen(url).read()
            data1 = data.decode("utf-8")

            content = re.search('<meta name="description" content="', data1)
            start = content.end()
            end = start + 500
            content1 = data1[start:end]

            content = re.search("definition, ", content1)
            start = content.end()
            content2 = content1[start:]

            content = re.search(' See more', content2)
            end = content.start()
            content3 = content2[:end]   
            
            display_meaning = scrolledtext.ScrolledText(window7, font = ("Arial", 15), height=5, width=60, fg = "black", bg = "#FFE4B5")
            display_meaning.place(x = 750, y=400, anchor="center")
            display_meaning.insert('end', content3)
            display_meaning.config(state='disabled')
            
            def listen2():
                        
                x=content3
                engine = pyttsx3.init()
                voices = engine.getProperty('voices')
                engine.setProperty('voice', voices[1].id)
                engine.say(x)
                engine.runAndWait()
                
            v2 =tk.Button(window7, text="Listen to word", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 12), command=listen1).place(x=550, y=500)    
            v3 =tk.Button(window7, text="Listen to meaning", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 12), command=listen2).place(x=750, y=500)
    
        except:
            tk.messagebox.showwarning("Error", 'Sorry, We could not find results for the word "' + keyword + '" :( ')
        
        
    def listen1():
            
        x=texttk.get()
        engine = pyttsx3.init()
        voices = engine.getProperty('voices')
        engine.setProperty('voice', voices[1].id)
        engine.say(x)
        engine.runAndWait()
        
    l_1 = tk.Label(window7, text="Enter Word: ",font=("Arial", 25),bg="#FFE4B5").place(x=380, y=250)
    
    texttk = tk.Entry(window7, fg='blue', bg='white',borderwidth=7,font=("Arial", 20), width=30)
    texttk.place(x=570, y=253)
    
    v1 = tk.Button(window7, text="Search", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 18), command=dictionary).place(x=1045, y=246)
    
    
    back3_button = tk.Button(text="Back", bg="blue", fg="white", height=1, width=10, borderwidth=8, cursor="hand2",font=("Arial", 12), command=back3)
    back3_button.place(x=650,y=650)
    
    window7.mainloop()


# In[10]:


#done
def spell():
    
    global window11
    window11 = tk.Tk()
    window11.geometry("1530x795+0+0")
    window11.configure(bg="#FFE4B5")
    window11.title("Notes O2O")
    
    
    head = tk.Label(window11, text="\nSpell Check\n", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5").pack()
                
    def back3() :
        window11.destroy()
        others()

    def check():
        w=texttk.get()
        words=[]
        words.append(w)
        corrected_words = []
        result=[]
        
        

        for i in words:
            corrected_words.append(TextBlob(i))
           
        for i in corrected_words:
            z=i.correct()
            result.append(z)

            
        display_meaning = scrolledtext.ScrolledText(window11, font = ("Arial", 15), height=5, width=60, fg = "black", bg = "#FFE4B5")
        display_meaning.place(x = 750, y=400, anchor="center")
        display_meaning.insert('end', result)
        display_meaning.config(state='disabled')
            
            

        
    l_1 = tk.Label(window11, text="Enter Word: ",font=("Arial", 25),bg="#FFE4B5").place(x=380, y=250)
    
    texttk = tk.Entry(window11, fg='blue', bg='white',borderwidth=7,font=("Arial", 20), width=30)
    texttk.place(x=570, y=253)
    
    v1 = tk.Button(window11, text="Check", bg="green", fg="white", borderwidth=8, cursor="hand2",font=("Arial", 18), command=check).place(x=1045, y=246)
    
    
    back3_button = tk.Button(text="Back", bg="blue", fg="white", height=1, width=10, borderwidth=8, cursor="hand2",font=("Arial", 12), command=back3)
    back3_button.place(x=650,y=650)
    
    window11.mainloop()


# In[11]:


#done
def saa():
    
    global window12
    window12 = tk.Tk()
    window12.geometry("1530x795+0+0")
    window12.configure(bg="#FFE4B5")
    window12.title("Notes O2O")
    
    
    head = tk.Label(window12, text="\nSynonyms and Antonyms\n", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5").pack()
                
    def back3() :
        window12.destroy()
        others()

    def syn():
        
        word=texttk1.get()
        
        synonyms = []
        
        for syn in wordnet.synsets(word):
            for lm in syn.lemmas():
                 synonyms.append(lm.name())
           
        content1=set(synonyms)
        
        try:
           
            display_syn = scrolledtext.ScrolledText(window12, font = ("Arial", 15), height=4, width=60, fg = "black", bg = "#FFE4B5")
            display_syn.place(x = 750, y=340, anchor="center")
            display_syn.insert('end', content1)
            display_syn.config(state='disabled')
            
            
        except:
            tk.messagebox.showwarning("Error", 'Sorry, We could not find results for the word :( ')

    def ant():
        word=texttk2.get()
    
        antonyms = []

        for syn in wordnet.synsets(word):
            for lm in syn.lemmas():
                if lm.antonyms():
                    antonyms.append(lm.antonyms()[0].name())

        content2=set(antonyms)
        
        try: 
            
            display_ant = scrolledtext.ScrolledText(window12, font = ("Arial", 15), height=4, width=60, fg = "black", bg = "#FFE4B5")
            display_ant.place(x = 750, y=590, anchor="center")
            display_ant.insert('end', content2)
            display_ant.config(state='disabled')
            
            
        except:
            tk.messagebox.showwarning("Error", 'Sorry, We could not find results for the word :( ')
        
        
        
    l_1 = tk.Label(window12, text="Enter Word: ",font=("Arial", 20),bg="#FFE4B5").place(x=380, y=200)
    
    texttk1 = tk.Entry(window12, fg='blue', bg='white',borderwidth=5,font=("Arial", 18), width=30)
    texttk1.place(x=540, y=203)
    
    v1 = tk.Button(window12, text="Search Synonym", bg="green", fg="white", borderwidth=5, cursor="hand2",font=("Arial", 14), command=syn).place(x=960, y=199)
    
    
    l_2 = tk.Label(window12, text="Enter Word: ",font=("Arial", 20),bg="#FFE4B5").place(x=380, y=450)
    
    texttk2 = tk.Entry(window12, fg='blue', bg='white',borderwidth=5,font=("Arial", 18), width=30)
    texttk2.place(x=540, y=453)
    
    v1 = tk.Button(window12, text="Search Antonym", bg="green", fg="white", borderwidth=5, cursor="hand2",font=("Arial", 14), command=ant).place(x=960, y=449)
    
    
    back3_button = tk.Button(text="Back", bg="blue", fg="white", height=1, width=10, borderwidth=5, cursor="hand2",font=("Arial", 12), command=back3)
    back3_button.place(x=650,y=700)
    
    window12.mainloop()


# In[12]:


def minigame():
    import index
    index.start_main_page()
    others()


# In[13]:


#done
def feedback_tk():
    
    global window8
    window8 = tk.Tk()
    window8.geometry("1530x795+0+0")
    window8.configure(bg="#FFE4B5")
    window8.title("Notes O2O")
    
    
    head = tk.Label(window8, text="\nFeedback Form\n", font=("rockwell extra bold",45),fg="dark blue",bg="#FFE4B5").pack()
    
    
    def back3() :
        window8.destroy()
        others()


    def excel():

        # resize the width of columns in
        # excel spreadsheet
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 50

        # write given data to an excel spreadsheet
        # at particular location
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Email ID"
        sheet.cell(row=1, column=3).value = "Feedback"


    # Function to set focus (cursor)

    # Function to set focus
    def focus1(event):
        # set focus on the email_id_field box
        name_field.focus_set()
        
    def focus5(event):
        # set focus on the email_id_field box
        email_id_field.focus_set()


    # Function to set focus
    def focus6(event):
        # set focus on the Feedback_field box
        Feedback_field.focus_set()


    # Function for clearing the
    # contents of text entry boxes
    def clear():

        # clear the content of text entry box
        name_field.delete(0, END)
        email_id_field.delete(0, END)
        Feedback_field.delete(0, END)

        
    # Function to take data from GUI
    # window and write to an excel file
    def insert():

        # if user not fill any entry
        # then print "empty input"
        if (name_field.get() == "" and
            email_id_field.get() == "" and
            Feedback_field.get() == ""):

            print("empty input")

        else:

            # assigning the max row and max column
            # value upto which data is written
            # in an excel sheet to the variable
            current_row = sheet.max_row
            current_column = sheet.max_column

            # get method returns current text
            # as string which we write into
            # excel spreadsheet at particular location
            sheet.cell(row=current_row + 1, column=1).value = name_field.get()
            sheet.cell(row=current_row + 1, column=2).value = email_id_field.get()
            sheet.cell(row=current_row + 1, column=3).value = Feedback_field.get()

            # save the file
            wb.save('C:\\Users\\ts4kj\\Documents\\Jupyter Files\\TechSaksham\\Capstone\\excel.xlsx')

            # set focus on the name_field box
            name_field.focus_set()
            
            # call the clear() function
            clear()
            
            tk.messagebox.showinfo("Information", "Thankyou for your valuable feedback!!")
            back3()
    # globally declare wb and sheet variable
 
    # opening the existing excel file
    wb = load_workbook('C:\\Users\\ts4kj\\Documents\\Jupyter Files\\TechSaksham\\Capstone\\excel.xlsx')
 
    # create the sheet object
    sheet = wb.active
 

    # Driver code
        
    excel()
    name = tk.Label(window8, text="Name: ",font=("Arial", 20),bg="#FFE4B5").place(x=420, y=200)

    name_field = tk.Entry(window8, fg='blue', bg='white',borderwidth=7,font=("Arial", 20), width=30)
    name_field.place(x=600, y=200)

    email_id = tk.Label(window8, text="Email ID: ",font=("Arial", 20),bg="#FFE4B5").place(x=420, y=300)

    email_id_field = tk.Entry(window8, fg='blue', bg='white',borderwidth=7,font=("Arial", 20), width=30)
    email_id_field.place(x=600, y=300)

    Feedback = tk.Label(window8, text="Feedback: ",font=("Arial", 20),bg="#FFE4B5").place(x=420, y=400)

    Feedback_field = tk.Entry(window8, fg='blue', bg='white',borderwidth=7,font=("Arial", 20), width=30)
    Feedback_field.place(x=600, y=400)
        
    # bind method of widget is used for
    # the binding the function with the events
    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus6 function
    email_id_field.bind("<Return>", focus6)

    # call excel function
    excel()

    back3_button = tk.Button(text="Back", bg="blue", fg="white", height=1, width=10, borderwidth=8, cursor="hand2",font=("Arial", 12), command=back3)
    back3_button.place(x=500,y=600)

    submit_button = tk.Button(text="Submit", bg="green", fg="white", height=1, width=10, borderwidth=8, cursor="hand2",font=("Arial", 12), command=insert)
    submit_button.place(x=800,y=600)

    # start the GUI
    window8.mainloop()
        


# In[14]:


main_screen()


# In[ ]:





# In[ ]:





# In[ ]:




